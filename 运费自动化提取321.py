import os
import sys
import pandas as pd
from openpyxl import load_workbook

MUNICIPALITIES = {'北京市', '上海市', '天津市', '重庆市'}


# ======================================================
# 地址解析
# ======================================================

def parse_address(addr):

    addr = str(addr).strip()

    if addr[:3] in MUNICIPALITIES:
        province = addr[:3]
        pe = 3

    elif '自治区' in addr:
        pe = addr.index('自治区') + 3
        province = addr[:pe]

    elif '省' in addr:
        pe = addr.index('省') + 1
        province = addr[:pe]

    else:
        province = addr[:3]
        pe = 3

    try:
        ce = addr.index('市', pe) + 1
    except:
        ce = pe

    city = addr[pe:ce]

    if not city:
        city = province

    endings = ['区', '县', '镇', '乡', '旗']

    de = len(addr)

    for ch in endings:

        try:
            pos = addr.index(ch, ce) + 1

            if pos < de:
                de = pos

        except:
            pass

    try:
        pos = addr.index('市', ce) + 1

        if pos < de:
            de = pos

    except:
        pass

    district = addr[ce:de]
    detail = addr[de:]

    for prefix in [addr[:de], city + district, district]:

        if prefix and detail.startswith(prefix):
            detail = detail[len(prefix):]
            break

    return province, city, district, detail


# ======================================================
# 自动寻找文件
# ======================================================

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

# 12
path_12 = None

if os.path.exists(os.path.join(CURRENT_DIR, "12.xlsx")):
    path_12 = os.path.join(CURRENT_DIR, "12.xlsx")

# 123
path_123 = None

if os.path.exists(os.path.join(CURRENT_DIR, "123.xlsx")):
    path_123 = os.path.join(CURRENT_DIR, "123.xlsx")

# 检查
if not path_12 or not path_123:

    print("\n【错误】缺少xlsx文件！")
    print("请确保：")
    print("12.xlsx")
    print("123.xlsx")
    print("在同一个文件夹内")

    input("\n按回车退出...")
    sys.exit()

output_path = os.path.join(CURRENT_DIR, "123_output.xlsx")


# ======================================================
# 读取12
# ======================================================

print("=================== 第一步：读取12数据 ===================")

df_12 = pd.read_excel(path_12, dtype=str)

df_12.columns = df_12.columns.str.strip()

# 全部转字符串，防止SKU/手机号/单号精度丢失
df_12 = df_12.fillna('')

# ------------------------------------------------------
# 自动将所有“广西壮族自治区”替换为“广西省”
# ------------------------------------------------------
if '收货地址' in df_12.columns:
    df_12['收货地址'] = df_12['收货地址'].str.replace('广西壮族自治区', '广西省')
    print("已将源数据中的'广西壮族自治区'全局替换为'广西省'")

print("12表字段：")
print(list(df_12.columns))


# ======================================================
# 打开123模板并重构列位置（精准寻找目标并插入左侧）
# ======================================================

print("\n=================== 第二步：重构123模板列位置 ===================")

wb = load_workbook(path_123)
ws = wb.active

# 1. 扫描当前所有表头
current_headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        current_headers[str(val).strip()] = col

# 2. 寻找锚点列（优先兼容你手动改好的'收货地址（复制）'，备用寻找原'收货地址'）
anchor_idx = current_headers.get('收货地址（复制）') 
if not anchor_idx:
    anchor_idx = current_headers.get('收货地址')

if anchor_idx:
    print(f"检测到锚点列在第 {anchor_idx} 列，准备进行结构调整...")
    
    # 强制确保该锚点列名为“收货地址（复制）”
    ws.cell(1, anchor_idx).value = '收货地址（复制）'
    
    # 检查左侧是否已经存在需要的解析列（防重复运行导致无限插入）
    if '收货省份' not in current_headers:
        print(f"正在原列（第 {anchor_idx} 列）的正左侧自动插入 4 个解析列...")
        
        # 在锚点位置插入4列，原本的“收货地址（复制）”及其右侧列会自动往右平移4格，公式也会自动适应
        ws.insert_cols(anchor_idx, 4)
        
        # 给新插入的 4 列写上表头
        ws.cell(1, anchor_idx).value = '收货省份'
        ws.cell(1, anchor_idx + 1).value = '收货城市'
        ws.cell(1, anchor_idx + 2).value = '收货区县'
        ws.cell(1, anchor_idx + 3).value = '收货地址' # 原“详细地址”改名为“收货地址”
        
        print("【成功】4个解析列已精确插入至左侧！")
    else:
        print("【提示】解析列已存在，跳过插入操作。")
else:
    print("【警告】未在123表中检测到 '收货地址' 或 '收货地址（复制）' 列！新列将在最右侧追加。")

# 3. 重新获取最新的表头字典（因为插入列会导致原本的列号全部发生变化，必须重新映射）
headers_123 = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header:
        headers_123[str(header).strip()] = col

print("\n更新后的123表字段：")
print(list(headers_123.keys()))


# ======================================================
# 行数校验与多余行清理
# ======================================================

excel_data_rows = ws.max_row - 1
target_data_count = len(df_12)

print(f"\n[行数统计] 12表实际数据：{target_data_count} 行 | 123模板检测到：{excel_data_rows} 行")

if excel_data_rows > target_data_count:
    start_delete_row = target_data_count + 2
    rows_to_delete = excel_data_rows - target_data_count
    print(f"⚠️ [提示] 模板旧数据较多，正在自动清理末尾第 {start_delete_row} 行起的 {rows_to_delete} 行多余内容...")
    ws.delete_rows(idx=start_delete_row, amount=rows_to_delete)


# ======================================================
# 自动复制字段
# ======================================================

print("\n=================== 第三步：复制字段 ===================")

copy_map = {
    # 基础信息
    '其它出库业务单号': '其它出库业务单号',
    '收货人': '收货人',
    '收货电话': '收货电话',
    '收货地址': '收货地址（复制）', # 映射至改名后的列

    # 业务字段
    'SKU采购总⾦额（含税）': '单价',
    '采购数量（采购单位）': '数量',
    'SKU编码': 'SKU编码',
}

for source_col, target_col in copy_map.items():

    if source_col not in df_12.columns:
        print(f"【警告】12缺少字段：{source_col}")
        continue

    if target_col not in headers_123:
        print(f"【警告】123缺少字段：{target_col}")
        continue

    target_excel_col = headers_123[target_col]

    for i in range(target_data_count):
        excel_row = i + 2
        value = df_12.iloc[i][source_col]

        # 编码类字段强制字符串
        if target_col in ['SKU编码', '其它出库业务单号', '收货电话']:
            value = str(value).strip()
            if value.endswith('.0'):
                value = value[:-2]

        ws.cell(excel_row, target_excel_col).value = value

    print(f"【成功】{source_col} -> {target_col}")


# ======================================================
# 地址解析与备注写入
# ======================================================

print("\n=================== 第四步：地址解析与备注生成 ===================")

if '收货地址' in df_12.columns:

    parsed_data = []

    for addr in df_12['收货地址']:
        if str(addr).strip():
            parsed_data.append(parse_address(addr))
        else:
            parsed_data.append(("", "", "", ""))

    parsed_df = pd.DataFrame(
        parsed_data,
        columns=[
            '收货省份',
            '收货城市',
            '收货区县',
            '收货地址'  
        ]
    )

    # 备注拼接
    clean_addresses = (
        df_12['收货地址']
        .astype(str)
        .fillna('')
        .str.replace('nan', '')
        .str.strip()
    )

    remark_series = (
        df_12['收货人'].astype(str).fillna('').str.replace('nan', '')
        + " "
        + df_12['收货电话'].astype(str).fillna('').str.replace('nan', '')
        + " "
        + clean_addresses
    ).str.strip()

    parsed_df['备注'] = remark_series

    # 写入数据
    for field in parsed_df.columns:
        # 如果不存在则自动新增（主要针对“备注”列）
        if field not in headers_123:
            new_col = ws.max_column + 1
            ws.cell(1, new_col).value = field
            headers_123[field] = new_col
            print(f"【自动新增列追加至右侧】{field}")

        target_excel_col = headers_123[field]

        for i in range(target_data_count):
            excel_row = i + 2
            value = parsed_df.iloc[i][field]
            ws.cell(excel_row, target_excel_col).value = value

        print(f"【成功】写入：{field}")


# ======================================================
# 保存
# ======================================================

print("\n=================== 第五步：保存文件 ===================")

wb.save(output_path)

print("\n【✨ 全部完成 ✨】")
print("列结构已精准调整：4个解析列已精确插入至 '收货地址（复制）' 列的左侧")
print("已全局替换：广西壮族自治区 -> 广西省")
print(f"12 表的 {target_data_count} 行数据已全部完整写入！多余旧数据已清理。")
print(f"输出文件：{output_path}")

input("\n按回车退出...")
